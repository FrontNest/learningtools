"""
Teszt Excel fajl generalasa a ber tablahoz.
A kepernyon latott adatsor alapjan.

Futtatas:
  cd c:\\Users\\lol\\______ORSI\\GitHub\\FrontnestStudio\\solutions\\tools\\tvSearch\\jogszabalyok
  python generate_test_data.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

WORKDIR = Path(__file__).parent

# -------------------------------------------------------------------------
# BER tabla adatsorok a kepernyon latott mintanak megfeleloen
# Fejlecek: sorszam | tol | ig | munkaltato | jogcim |
#           rendszeres jövedelem | nem rendszeres jövedelem |
#           ellatas_1 (gyes/gyed/tgyas/csed) | ellatas_2 (gyap/munkanelkuli/stb) |
#           befizetett nyugdijjarulek
# -------------------------------------------------------------------------
BER_HEADERS = [
    "sorszam",
    "tol",
    "ig",
    "munkaltato",
    "jogcim",
    "rendszeres_jovedelem",
    "nem_rendszeres_jovedelem",
    "ellatas_1",   # gyes/gyed/tgyas/csed
    "ellatas_2",   # gyap/munkanelkuli/stb
    "befizetett_nyugdijjarulek",
]

BER_ROWS = [
    #  #   tol           ig            munkaltato          jogcim                     rend       nem_rend   ell1   ell2   jar
    [ 1, "1986.01.01", "1986.12.31", "Cég 1",            "munkaviszony",            345678,      0,       6000,   None,  365  ],
    [ 2, "1986.03.15", "1986.05.30", None,               "fizetés nélküli szabadság",     0,    0,       None,   None,  None ],
    [ 3, "1986.06.01", "1986.06.15", None,               "táppénz",                       0,    0,       None,   None,  None ],
    [ 4, "1986.06.16", "1989.06.16", None,               "GYES",                          0,    0,       None,   5000,  342  ],
    [ 5, "1987.01.01", "1987.12.31", "Cég 2",            "munkaviszony",            456789,      0,       None,   None,  350  ],
    [ 6, "1988.01.01", "1988.01.30", "M1",               "megbízásos",                   15,    0,       None,   None,  400  ],
    [ 7, "1988.02.01", "1988.03.30", "M2",               "megbízásos",          123456789,       0,       None,   None,  410  ],
    [ 8, "2000.01.01", "2000.12.01", "Cég 3",            "munkaviszony",            839265,      0,       None,   None,    4  ],
    [ 9, "2000.03.03", "2000.12.31", "M3",               "megbízásos",                  650,    0,       None,   None,    5  ],
    [10, "2000.03.04", "2001.01.01", "M4",               "megbízásos",            12345678,      0,       None,   None,    6  ],
    [11, "2000.03.05", "2001.01.02", "M5",               "megbízásos",            12345678,      0,       None,   None,    7  ],
    [12, "2000.03.06", "2001.01.03", "egyéni vállalkozó","egyéni vállalkozó főállás", 12345678,  0,       None,   None,    8  ],
    # FNYSZ 30-napos szabaly tesztesehez: 1990-ben 45 napos FNYSZ
    [13, "1990.01.01", "1990.02.14", None,               "fizetés nélküli szabadság",       0,  0,       None,   None, None  ],
]

# -------------------------------------------------------------------------
# SZOLGALATI IDO tabla adatsorok (a ber soroknak megfeleloen)
# A gyakorlatban ez is a nyers NYUFIG adatbol szarmazik
# Fejlecek: sorszam | tol | ig | inclusive_napok | munkaltato | jogcim | minosites
# -------------------------------------------------------------------------
SZOLG_HEADERS = [
    "sorszam",
    "tol",
    "ig",
    "inclusive_napok",
    "munkaltato",
    "jogcim",
    "minosites",
]

SZOLG_ROWS = [
    #  #   tol           ig            nap   munkaltato          jogcim                         min
    [ 1, "1986.01.01", "1986.12.31",  365, "Cég 1",            "munkaviszony",                "N"  ],
    [ 2, "1986.03.15", "1986.05.30",   77, None,               "fizetés nélküli szabadság",   "N"  ],
    [ 3, "1986.06.01", "1986.06.15",   15, None,               "táppénz",                     "N"  ],
    [ 4, "1986.06.16", "1989.06.16", 1097, None,               "GYES",                        "G"  ],
    [ 5, "1987.01.01", "1987.12.31",  365, "Cég 2",            "munkaviszony",                "N"  ],
    [ 6, "1988.01.01", "1988.01.30",   30, "M1",               "megbízásos",                  "N"  ],
    [ 7, "1988.02.01", "1988.03.30",   59, "M2",               "megbízásos",                  "N"  ],
    [ 8, "2000.01.01", "2000.12.01",  336, "Cég 3",            "munkaviszony",                "N"  ],
    [ 9, "2000.03.03", "2000.12.31",  304, "M3",               "megbízásos",                  "N"  ],
    [10, "2000.03.04", "2001.01.01",  304, "M4",               "megbízásos",                  "N"  ],
    [11, "2000.03.05", "2001.01.02",  304, "M5",               "megbízásos",                  "N"  ],
    [12, "2000.03.06", "2001.01.03",  304, "egyéni vállalkozó","egyéni vállalkozó főállás",   "J"  ],
    # FNYSZ 30-napos szabaly tesztesehez: 1990-ben 45 napos FNYSZ, nincs mogotte munkaviszony
    [13, "1990.01.01", "1990.02.14",   45, None,               "fizetés nélküli szabadság",   "N"  ],
]


def style_header_row(ws, row_idx: int, fill_color: str = "4472C4"):
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align


def auto_column_widths(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


def write_sheet(wb: openpyxl.Workbook, sheet_name: str, headers: list, rows: list):
    ws = wb.create_sheet(title=sheet_name)
    ws.freeze_panes = "A2"
    ws.append(headers)
    style_header_row(ws, 1)
    for row in rows:
        ws.append(row)
    auto_column_widths(ws)
    return ws


def main():
    output_path = WORKDIR / "test_input.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    write_sheet(wb, "ber", BER_HEADERS, BER_ROWS)
    write_sheet(wb, "szolgalati_ido", SZOLG_HEADERS, SZOLG_ROWS)

    wb.save(output_path)
    print(f"Test Excel keszult: {output_path}")


if __name__ == "__main__":
    main()
